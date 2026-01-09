import os
from typing import Dict, Any, List, Union
import openpyxl
from openpyxl.utils import get_column_letter

from core.action_base import ActionBase

# Global session storage for open Excel workbooks
# Key: alias, Value: {"wb": workbook_obj, "path": file_path}
EXCEL_SESSIONS = {}

class OpenExcelAction(ActionBase):
    @property
    def name(self) -> str:
        return "打开 Excel"
    
    @property
    def description(self) -> str:
        return "打开一个 Excel 文件并保持会话。"
    
    def get_param_schema(self) -> List[Dict[str, Any]]:
        return [
            {"name": "file_path", "label": "文件路径", "type": "string", "ui_options": {"browse_type": "file"}},
            {"name": "alias", "label": "会话别名", "type": "string", "default": "default", "variable_type": "Excel对象"},
            {"name": "data_only", "label": "只读取数值(忽略公式)", "type": "bool", "default": True},
            {"name": "read_only", "label": "只读模式", "type": "bool", "default": False}
        ]
    
    def execute(self, context: Dict[str, Any]) -> bool:
        file_path = self.params.get("file_path")
        alias = self.params.get("alias", "default")
        data_only = self.params.get("data_only", True)
        read_only = self.params.get("read_only", False)
        
        try:
            # Format file path with context if needed
            file_path = file_path.format(**context)
        except:
            pass
            
        if not file_path or not os.path.exists(file_path):
            print(f"[OpenExcel] File not found: {file_path}")
            return False
            
        try:
            print(f"[OpenExcel] Opening {file_path} (Alias: {alias})...")
            wb = openpyxl.load_workbook(filename=file_path, data_only=data_only, read_only=read_only)
            EXCEL_SESSIONS[alias] = {"wb": wb, "path": file_path, "read_only": read_only}
            return True
        except Exception as e:
            print(f"[OpenExcel] Error: {e}")
            return False

class ReadExcelAction(ActionBase):
    @property
    def name(self) -> str:
        return "读取 Excel"
    
    @property
    def description(self) -> str:
        return "读取 Excel 的 Sheet、单元格或区域。"
    
    def get_param_schema(self) -> List[Dict[str, Any]]:
        return [
            {"name": "alias", "label": "会话别名", "type": "string", "default": "default", "variable_type": "Excel对象"},
            {"name": "sheet_name", "label": "Sheet 名称 (空则当前)", "type": "string", "default": "", "advanced": True},
            {"name": "read_type", "label": "读取类型", "type": "string", "default": "Cell", "options": ["Cell", "Range", "Sheet"]},
            {"name": "address", "label": "地址 (如 A1 或 A1:B2)", "type": "string", "default": "A1"},
            {"name": "output_variable", "label": "保存到变量", "type": "string", "default": "excel_data", "variable_type": "一般变量"}
        ]
    
    def execute(self, context: Dict[str, Any]) -> bool:
        alias = self.params.get("alias", "default")
        sheet_name = self.params.get("sheet_name")
        read_type = self.params.get("read_type", "Cell") # Cell, Range, Sheet
        address = self.params.get("address", "A1")
        output_var = self.params.get("output_variable", "excel_data")
        
        if alias not in EXCEL_SESSIONS:
            print(f"[ReadExcel] Session '{alias}' not found. Did you Open Excel?")
            return False
            
        session = EXCEL_SESSIONS[alias]
        wb = session["wb"]
        
        try:
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    print(f"[ReadExcel] Sheet '{sheet_name}' not found.")
                    return False
            else:
                ws = wb.active
            
            result = None
            
            if read_type == "Cell":
                # Support A1
                result = ws[address].value
                print(f"[ReadExcel] Cell {address} = {result}")
                
            elif read_type == "Range":
                # Support A1:B2 -> list of lists
                data = []
                # ws[address] returns a tuple of tuples of cells
                rows = ws[address]
                if not isinstance(rows, tuple): # Single cell fallback
                     rows = ((rows,),)
                
                for row in rows:
                    row_data = [cell.value for cell in row]
                    data.append(row_data)
                result = data
                print(f"[ReadExcel] Range {address} read {len(data)} rows.")
                
            elif read_type == "Sheet":
                # Read all used cells
                data = []
                # iter_rows might be slow for huge sheets in read-only mode?
                # openpyxl is generally okay.
                headers = []
                is_first = True
                for row in ws.iter_rows(values_only=True):
                    if is_first:
                        headers = list(row)
                        is_first = False
                        continue
                    
                    # Create dict based on headers
                    item = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and headers[i] is not None:
                            item[str(headers[i])] = val
                    data.append(item)
                result = data
                print(f"[ReadExcel] Sheet read {len(data)} rows.")
            
            context[output_var] = result
            return True
            
        except Exception as e:
            print(f"[ReadExcel] Error: {e}")
            return False

class GetExcelRowCountAction(ActionBase):
    @property
    def name(self) -> str:
        return "获取 Excel 总行数"
    
    @property
    def description(self) -> str:
        return "获取指定 Sheet 的总行数。"
    
    def get_param_schema(self) -> List[Dict[str, Any]]:
        return [
            {"name": "alias", "label": "会话别名", "type": "string", "default": "default", "variable_type": "Excel对象"},
            {"name": "sheet_name", "label": "Sheet 名称 (空则使用当前)", "type": "string", "default": ""},
            {"name": "output_variable", "label": "保存到变量", "type": "string", "default": "row_count", "variable_type": "一般变量"}
        ]
    
    def execute(self, context: Dict[str, Any]) -> bool:
        alias = self.params.get("alias", "default")
        sheet_name = self.params.get("sheet_name")
        output_var = self.params.get("output_variable", "row_count")
        
        if alias not in EXCEL_SESSIONS:
            print(f"[GetExcelRowCount] Session '{alias}' not found.")
            return False
            
        wb = EXCEL_SESSIONS[alias]["wb"]
        
        try:
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    print(f"[GetExcelRowCount] Sheet '{sheet_name}' not found.")
                    return False
            else:
                ws = wb.active
                
            count = ws.max_row
            context[output_var] = count
            print(f"[GetExcelRowCount] Count: {count}")
            return True
        except Exception as e:
            print(f"[GetExcelRowCount] Error: {e}")
            return False

class WriteExcelAction(ActionBase):
    @property
    def name(self) -> str:
        return "写入 Excel"
    
    @property
    def description(self) -> str:
        return "写入数据到 Excel 的单元格或区域。"
    
    def get_param_schema(self) -> List[Dict[str, Any]]:
        return [
            {"name": "alias", "label": "会话别名", "type": "string", "default": "default", "variable_type": "Excel对象"},
            {"name": "sheet_name", "label": "Sheet 名称 (空则当前)", "type": "string", "default": "", "advanced": True},
            {"name": "write_type", "label": "写入类型", "type": "string", "default": "Cell", "options": ["Cell", "Range"]},
            {"name": "address", "label": "地址 (如 A1)", "type": "string", "default": "A1"},
            {"name": "value", "label": "写入值 (支持变量 {var})", "type": "string", "default": ""}
        ]
    
    def execute(self, context: Dict[str, Any]) -> bool:
        alias = self.params.get("alias", "default")
        sheet_name = self.params.get("sheet_name")
        write_type = self.params.get("write_type", "Cell")
        address = self.params.get("address", "A1")
        value_raw = self.params.get("value", "")
        
        if alias not in EXCEL_SESSIONS:
            print(f"[WriteExcel] Session '{alias}' not found.")
            return False
            
        session = EXCEL_SESSIONS[alias]
        if session.get("read_only"):
             print(f"[WriteExcel] Session '{alias}' is read-only.")
             return False
             
        wb = session["wb"]
        
        try:
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    # Create if not exists? Or fail? Fail is safer.
                    print(f"[WriteExcel] Sheet '{sheet_name}' not found.")
                    return False
            else:
                ws = wb.active
                
            # Resolve value from context if it looks like a variable
            # or if it's a complex object (list/dict) passed via context reference?
            # The input is string. If user wants to pass a list (for Range), they might need to use SetVariable with a list
            # and then reference it here?
            # But "value" param is string.
            # If value starts with $ or similar, we could treat it as var lookup.
            # Or use standard format syntax.
            
            value_to_write = value_raw
            
            # Try to resolve variable or parse JSON
            if isinstance(value_raw, str):
                # 1. Check for direct variable reference like "{var_name}"
                if value_raw.startswith("{") and value_raw.endswith("}") and value_raw.count("{") == 1:
                    var_name = value_raw[1:-1]
                    if var_name in context:
                        value_to_write = context[var_name]
                    else:
                        # Fallback to formatting if var not found (though unlikely to work for list)
                         try:
                            value_to_write = value_raw.format(**context)
                         except:
                            pass
                else:
                    # 2. Try JSON parsing for lists/dicts
                    import json
                    try:
                        value_to_write = json.loads(value_raw)
                    except:
                        # 3. Fallback to string formatting
                        try:
                            value_to_write = value_raw.format(**context)
                        except:
                            pass
            
            if write_type == "Cell":
                ws[address] = value_to_write
                print(f"[WriteExcel] Wrote to {address}")
                
            elif write_type == "Range":
                # Expecting list of lists
                start_cell = ws[address]
                start_row = start_cell.row
                start_col = start_cell.column
                
                if isinstance(value_to_write, list):
                    for r_idx, row_data in enumerate(value_to_write):
                        if isinstance(row_data, list):
                            for c_idx, val in enumerate(row_data):
                                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
                        else:
                            # Single list (one row)
                             ws.cell(row=start_row, column=start_col + r_idx, value=row_data)
                    print(f"[WriteExcel] Wrote range starting at {address}")
                else:
                    print("[WriteExcel] Value for Range must be a list.")
                    return False
                    
            return True
        except Exception as e:
            print(f"[WriteExcel] Error: {e}")
            return False

class SaveExcelAction(ActionBase):
    @property
    def name(self) -> str:
        return "保存 Excel"
    
    @property
    def description(self) -> str:
        return "保存 Excel 工作簿到文件。"
    
    def get_param_schema(self) -> List[Dict[str, Any]]:
        return [
            {"name": "alias", "label": "会话别名", "type": "string", "default": "default", "variable_type": "Excel对象"},
            {"name": "file_path", "label": "另存为路径 (空则覆盖原文件)", "type": "string", "default": "", "ui_options": {"browse_type": "file", "file_filter": "Excel Files (*.xlsx *.xls)"}}
        ]
    
    def execute(self, context: Dict[str, Any]) -> bool:
        alias = self.params.get("alias", "default")
        new_path = self.params.get("file_path")
        
        if alias not in EXCEL_SESSIONS:
            print(f"[SaveExcel] Session '{alias}' not found.")
            return False
            
        session = EXCEL_SESSIONS[alias]
        wb = session["wb"]
        original_path = session["path"]
        
        save_path = new_path if new_path else original_path
        
        try:
            wb.save(save_path)
            print(f"[SaveExcel] Saved to {save_path}")
            return True
        except Exception as e:
            print(f"[SaveExcel] Error: {e}")
            return False

class CloseExcelAction(ActionBase):
    @property
    def name(self) -> str:
        return "关闭 Excel"
    
    @property
    def description(self) -> str:
        return "保存并关闭 Excel 会话。"
    
    def get_param_schema(self) -> List[Dict[str, Any]]:
        return [
            {"name": "alias", "label": "会话别名", "type": "string", "default": "default", "variable_type": "Excel对象"},
            {"name": "save", "label": "是否保存", "type": "bool", "default": True}
        ]
    
    def execute(self, context: Dict[str, Any]) -> bool:
        alias = self.params.get("alias", "default")
        save = self.params.get("save", True)
        
        if alias not in EXCEL_SESSIONS:
            return True
            
        session = EXCEL_SESSIONS[alias]
        wb = session["wb"]
        path = session["path"]
        
        try:
            if save and not session.get("read_only"):
                print(f"[CloseExcel] Saving {path}...")
                wb.save(path)
            
            wb.close()
            del EXCEL_SESSIONS[alias]
            print(f"[CloseExcel] Closed session '{alias}'")
            return True
        except Exception as e:
            print(f"[CloseExcel] Error: {e}")
            return False
